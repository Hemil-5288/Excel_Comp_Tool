# save as app.py (replace your existing backend file)
import io
import re
import json
from copy import copy
from datetime import datetime
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = FastAPI(title="Excel Comparison API", version="1.0.0")

from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

NUMERIC_TOLERANCE = 1

def coerce_numeric(value):
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    s = re.sub(r"[^\d\.\-]+", "", s)
    try:
        return float(s)
    except Exception:
        return None

def to_int_or_none(value):
    num = coerce_numeric(value)
    if num is None:
        return None
    try:
        return int(num)
    except Exception:
        return None

def norm_text(value):
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()

def values_equal(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    oi = to_int_or_none(orig_val)
    wi = to_int_or_none(web_val)
    if oi is not None and wi is not None:
        return abs(wi - oi) <= tol
    return norm_text(orig_val) == norm_text(web_val)

def values_different(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    return not values_equal(orig_val, web_val, tol=tol)

def row_key(row, max_parts=2):
    key_parts = []
    for col in row.index:
        iv = to_int_or_none(row[col])
        if iv is not None:
            key_parts.append(("num", iv))
        else:
            sv = norm_text(row[col])
            if sv != "":
                key_parts.append(("str", sv.lower()))
        if len(key_parts) >= max_parts:
            break
    return tuple(key_parts)

def build_key_index(df, max_parts=2):
    idx = {}
    for i, row in df.iterrows():
        k = row_key(row, max_parts=max_parts)
        idx.setdefault(k, []).append(i)
    return idx

def safe_copy_cell_style(source_cell, target_cell):
    try:
        if source_cell is None or target_cell is None:
            return
        if hasattr(source_cell, "has_style") and source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass

def safe_get_column_width(worksheet, col_letter):
    try:
        if col_letter in worksheet.column_dimensions:
            return worksheet.column_dimensions[col_letter].width
        return None
    except Exception:
        return None

def norm_for_json(v):
    """Normalize a cell for JSON output but preserve numeric types and leading-zero strings."""
    if pd.isna(v) or v is None:
        return ""
    # Keep true numeric types as numbers
    if isinstance(v, (int, float, bool)):
        # convert numpy types to python scalars if needed
        try:
            return int(v) if isinstance(v, int) and not isinstance(v, bool) else float(v) if isinstance(v, float) else v
        except Exception:
            return v
    # everything else becomes string (preserves leading zeros)
    return str(v)

# ---------- CORE compare function (returns both workbook bytes and structured JSON sheets) ----------
def compare_excel_with_gain_summary_inline(
    original_bytes: bytes,
    website_bytes: bytes,
    sheets_config: Optional[Dict[str, Dict[str, int]]] = None,
) -> Dict[str, Any]:

    if sheets_config is None:
        sheets_config = {
            'Gain Summary': {'header_row': 2, 'data_start_row': 3},
            '8938':         {'header_row': 6, 'data_start_row': 7},
            'FBAR':         {'header_row': 2, 'data_start_row': 3},
        }

    red_fill   = PatternFill(start_color="fbd9d3", end_color="fbd9d3", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    original_wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    website_wb  = openpyxl.load_workbook(io.BytesIO(website_bytes))
    output_wb   = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])

    summary_rows = [["Sheet", "Rows Compared", "Cells Different", "Common Rows", "Only in Original", "Only in Website"]]

    # This dict will be the single source-of-truth for sheet row categories (JSON-serializable)
    sheets_structured: Dict[str, Dict[str, Any]] = {}

    def acct_key(v):
        if pd.isna(v) or v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        num = to_int_or_none(s)
        if num is not None:
            return str(num)
        return s.lower()

    for sheet_name, cfg in sheets_config.items():
        try:
            # read with pandas (will raise if sheet not present)
            df_orig = pd.read_excel(io.BytesIO(original_bytes), sheet_name=sheet_name, header=cfg['header_row'] - 1)
            df_web  = pd.read_excel(io.BytesIO(website_bytes),  sheet_name=sheet_name, header=cfg['header_row'] - 1)

            common_cols = [c for c in df_orig.columns if c in df_web.columns]

            # Prepare structured lists (each begins with header row)
            structured_common = []
            structured_diff   = []
            structured_only_orig = []
            structured_only_web  = []

            if not common_cols:
                # create empty sheets (still create workbook sheets for download)
                output_wb.create_sheet(title=sheet_name)
                output_wb.create_sheet(title=f"{sheet_name} Common Rows")
                summary_rows.append([sheet_name, 0, 0, 0, 0, 0])
                sheets_structured[sheet_name] = {
                    "common_rows": structured_common,
                    "different_rows": structured_diff,
                    "only_in_original_rows": structured_only_orig,
                    "only_in_website_rows": structured_only_web,
                    "rows_compared": 0,
                    "cells_different": 0
                }
                continue

            # shrink to common columns
            df_orig = df_orig[common_cols].copy()
            df_web  = df_web[common_cols].copy()

            original_ws = original_wb[sheet_name] if sheet_name in original_wb.sheetnames else None
            website_ws  = website_wb[sheet_name]  if sheet_name in website_wb.sheetnames  else None
            main_ws     = output_wb.create_sheet(title=sheet_name)

            diff_count       = 0
            rows_compared    = 0
            common_rows_list = []

            # ---------- Helper: write header for structured lists ----------
            def make_tripled_headers():
                headers = []
                for col in common_cols:
                    headers.extend([f"{col} (Original)", f"{col} (Website)", f"{col} (Diff)"])
                # set header row to structured lists
                structured_common.append([h for h in headers])
                structured_diff.append([h for h in headers])
                structured_only_orig.append([h for h in headers])
                structured_only_web.append([h for h in headers])
                # write headers to workbook
                for ci, h in enumerate(headers, start=1):
                    main_ws.cell(row=1, column=ci, value=h)
                    if original_ws is not None:
                        try:
                            # best-effort style copy
                            src = original_ws.cell(row=cfg['header_row'], column=((ci - 1) // 3) + 1)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            pass
                # set widths
                for idx_col, _ in enumerate(common_cols, start=1):
                    width = None
                    if original_ws is not None:
                        width = safe_get_column_width(original_ws, get_column_letter(idx_col))
                    for offset in range(3):
                        out_col_letter = get_column_letter((idx_col - 1) * 3 + offset + 1)
                        if width:
                            main_ws.column_dimensions[out_col_letter].width = width

            def make_simple_headers():
                headers = [c for c in common_cols] + ["Row Status"]
                structured_common.append([h for h in headers])
                structured_diff.append([h for h in headers])
                structured_only_orig.append([h for h in headers])
                structured_only_web.append([h for h in headers])
                # write headers to workbook
                for ci, h in enumerate(headers, start=1):
                    main_ws.cell(row=1, column=ci, value=h)
                    if original_ws is not None:
                        try:
                            src = original_ws.cell(row=cfg['header_row'], column=ci)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            pass
                # width for status
                main_ws.column_dimensions[get_column_letter(len(headers))].width = 18

            # ---------- Per-sheet logic (write workbook AND build structured lists) ----------
            if sheet_name in ("Gain Summary", "8938", "FBAR") and ("Account Number" in df_orig.columns) and ("Account Number" in df_web.columns):
                # Use the tripled header format
                make_tripled_headers()

                # special handling: dedupe by account number
                df_orig = df_orig.drop_duplicates(subset=["Account Number"], keep="first").reset_index(drop=True)
                df_web  = df_web.drop_duplicates(subset=["Account Number"], keep="first").reset_index(drop=True)

                # Build key maps
                if sheet_name == "Gain Summary":
                    web_key_index = build_key_index(df_web, max_parts=2)
                    orig_keys_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

                    out_row = 2
                    for orig_idx, orig_row in df_orig.iterrows():
                        k = row_key(orig_row, max_parts=2)
                        match_indices = web_key_index.get(k, [])
                        match_idx = match_indices[0] if match_indices else None

                        if match_idx is None:
                            # Only in original: tripled row with website empty and Diff marked
                            triple_row = []
                            for col in common_cols:
                                o_val = norm_for_json(orig_row[col])
                                triple_row.extend([o_val, "", "Only in Original"])
                                # workbook write
                                main_ws.cell(row=out_row, column=((common_cols.index(col)) * 3) + 1, value=o_val)
                                main_ws.cell(row=out_row, column=((common_cols.index(col)) * 3) + 3, value="Only in Original").fill = red_fill
                            structured_only_orig.append(triple_row)
                            out_row += 1
                            rows_compared += 1
                            diff_count += 1
                            continue

                        web_row = df_web.loc[match_idx]
                        # check all columns same
                        all_same = all(values_equal(orig_row[col], web_row[col]) for col in common_cols)
                        if all_same:
                            # For common rows add the tripled row but leave diff cells blank
                            triple_row = []
                            for col in common_cols:
                                val = norm_for_json(orig_row[col])
                                triple_row.extend([val, val, ""])
                                # workbook: put orig, site and blank diff
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            structured_common.append(triple_row)
                            out_row += 1
                            # do NOT increment diff_count
                            continue

                        # They differ: create tripled row with per-column diffs
                        row_diffs_here = 0
                        triple_row = []
                        for col in common_cols:
                            o_val = orig_row[col]
                            w_val = web_row[col]
                            is_diff = values_different(o_val, w_val)
                            diff_val = ""
                            if is_diff:
                                o_num = coerce_numeric(o_val)
                                w_num = coerce_numeric(w_val)
                                if o_num is not None and w_num is not None:
                                    # numeric diff
                                    diff_val = w_num - o_num
                                else:
                                    diff_val = "DIFF"
                                row_diffs_here += 1

                            triple_row.extend([norm_for_json(o_val), norm_for_json(w_val), norm_for_json(diff_val)])
                            # workbook write
                            ci = common_cols.index(col)
                            c_orig = main_ws.cell(row=out_row, column=ci * 3 + 1, value=o_val)
                            c_web  = main_ws.cell(row=out_row, column=ci * 3 + 2, value=w_val)
                            c_diff = main_ws.cell(row=out_row, column=ci * 3 + 3, value=diff_val)
                            if is_diff:
                                c_orig.fill = red_fill
                                c_web.fill  = green_fill
                                c_diff.fill = red_fill

                        structured_diff.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += row_diffs_here

                    # Now web-only rows
                    for web_idx, web_row in df_web.iterrows():
                        k = row_key(web_row, max_parts=2)
                        if k in orig_keys_set:
                            continue
                        triple_row = []
                        for col in common_cols:
                            w_val = norm_for_json(web_row[col])
                            triple_row.extend(["", w_val, "Only in Website"])
                        structured_only_web.append(triple_row)
                        # workbook write
                        for col in common_cols:
                            ci = common_cols.index(col)
                            main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Website").fill = green_fill
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1

                else:
                    # 8938 / FBAR: keyed by 'Account Number'
                    orig_map = {}
                    for i, r in df_orig.iterrows():
                        k = acct_key(r["Account Number"])
                        if k is not None and k not in orig_map:
                            orig_map[k] = i
                    web_map = {}
                    for i, r in df_web.iterrows():
                        k = acct_key(r["Account Number"])
                        if k is not None and k not in web_map:
                            web_map[k] = i

                    out_row = 2
                    for acc in orig_map.keys():
                        orig_idx = orig_map[acc]
                        orig_row = df_orig.loc[orig_idx]
                        if acc not in web_map:
                            triple_row = []
                            for col in common_cols:
                                o_val = norm_for_json(orig_row[col])
                                triple_row.extend([o_val, "", "Only in Original"])
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Original").fill = red_fill
                            structured_only_orig.append(triple_row)
                            out_row += 1
                            rows_compared += 1
                            diff_count += 1
                            continue

                        web_idx = web_map[acc]
                        web_row = df_web.loc[web_idx]
                        all_same = all(values_equal(orig_row[col], web_row[col]) for col in common_cols)
                        if all_same:
                            triple_row = []
                            for col in common_cols:
                                val = norm_for_json(orig_row[col])
                                triple_row.extend([val, val, ""])
                                ci = common_cols.index(col)
                                main_ws.cell(row=out_row, column=ci * 3 + 1, value=orig_row[col])
                                main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            structured_common.append(triple_row)
                            out_row += 1
                            continue

                        row_diffs_here = 0
                        triple_row = []
                        for col in common_cols:
                            o_val = orig_row[col]
                            w_val = web_row[col]
                            is_diff = values_different(o_val, w_val)
                            diff_val = ""
                            if is_diff:
                                o_num = coerce_numeric(o_val)
                                w_num = coerce_numeric(w_val)
                                if o_num is not None and w_num is not None:
                                    diff_val = w_num - o_num
                                else:
                                    diff_val = "DIFF"
                                row_diffs_here += 1

                            triple_row.extend([norm_for_json(o_val), norm_for_json(w_val), norm_for_json(diff_val)])
                            ci = common_cols.index(col)
                            c_orig = main_ws.cell(row=out_row, column=ci * 3 + 1, value=o_val)
                            c_web  = main_ws.cell(row=out_row, column=ci * 3 + 2, value=w_val)
                            c_diff = main_ws.cell(row=out_row, column=ci * 3 + 3, value=diff_val)
                            if is_diff:
                                c_orig.fill = red_fill
                                c_web.fill  = green_fill
                                c_diff.fill = red_fill

                        structured_diff.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += row_diffs_here

                    # web-only
                    for acc in web_map.keys():
                        if acc in orig_map:
                            continue
                        web_idx = web_map[acc]
                        web_row = df_web.loc[web_idx]
                        triple_row = []
                        for col in common_cols:
                            w_val = norm_for_json(web_row[col])
                            triple_row.extend(["", w_val, "Only in Website"])
                            ci = common_cols.index(col)
                            main_ws.cell(row=out_row, column=ci * 3 + 2, value=web_row[col])
                            main_ws.cell(row=out_row, column=ci * 3 + 3, value="Only in Website").fill = green_fill
                        structured_only_web.append(triple_row)
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1

            # Generic sheets (not the special triple format)
            else:
                # create simple header + status
                make_simple_headers()
                web_key_index = build_key_index(df_web, max_parts=2)
                out_row = 2

                # We'll also build a set of orig row keys to identify web-only rows easily
                orig_key_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

                for orig_idx, orig_row in df_orig.iterrows():
                    k = row_key(orig_row, max_parts=2)
                    match_indices = web_key_index.get(k, [])
                    match_idx = match_indices[0] if match_indices else None

                    if match_idx is None:
                        # Only in original
                        rowvals = [norm_for_json(orig_row[col]) for col in common_cols] + ["Only in Original"]
                        structured_only_orig.append(rowvals)
                        # workbook
                        for i, col in enumerate(common_cols, start=1):
                            main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Only in Original").fill = red_fill
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1
                        continue

                    web_row = df_web.loc[match_idx]
                    # if all equal => common
                    if all(values_equal(orig_row[col], web_row[col]) for col in common_cols):
                        rowvals = [norm_for_json(orig_row[col]) for col in common_cols] + ["Common"]
                        structured_common.append(rowvals)
                        # workbook
                        for i, col in enumerate(common_cols, start=1):
                            main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        out_row += 1
                        continue

                    # different
                    row_diffs_here = 0
                    rowvals = []
                    for col in common_cols:
                        o_val = orig_row[col]
                        w_val = web_row[col]
                        if values_different(o_val, w_val):
                            row_diffs_here += 1
                        rowvals.append(norm_for_json(o_val))
                    rowvals.append("Different")
                    structured_diff.append(rowvals)
                    # workbook
                    for i, col in enumerate(common_cols, start=1):
                        c = main_ws.cell(row=out_row, column=i, value=orig_row[col])
                        if values_different(orig_row[col], web_row[col]):
                            c.fill = red_fill
                    main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Different").fill = red_fill
                    out_row += 1
                    rows_compared += 1
                    diff_count += row_diffs_here

                # handle web-only rows
                for web_idx, web_row in df_web.iterrows():
                    k = row_key(web_row, max_parts=2)
                    if k in orig_key_set:
                        continue
                    rowvals = [norm_for_json(web_row[col]) for col in common_cols] + ["Only in Website"]
                    structured_only_web.append(rowvals)
                    # workbook
                    for i, col in enumerate(common_cols, start=1):
                        main_ws.cell(row=out_row, column=i, value=web_row[col])
                    main_ws.cell(row=out_row, column=len(common_cols) + 1, value="Only in Website").fill = green_fill
                    out_row += 1
                    rows_compared += 1
                    diff_count += 1

            # Auto-size columns (best-effort)
            for ws_auto in (main_ws,):
                for col in ws_auto.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            try:
                                max_len = max(max_len, len(str(cell.value)))
                            except Exception:
                                max_len = max(max_len, 10)
                    ws_auto.column_dimensions[col_letter].width = min(max_len + 2, 50)

            # Fill structured summary for this sheet
            sheets_structured[sheet_name] = {
                "common_rows": structured_common,
                "different_rows": structured_diff,
                "only_in_original_rows": structured_only_orig,
                "only_in_website_rows": structured_only_web,
                "rows_compared": rows_compared,
                "cells_different": diff_count
            }

            summary_rows.append([sheet_name, rows_compared, diff_count, len(structured_common), len(structured_only_orig), len(structured_only_web)])

        except Exception as e:
            # If an exception occurs, save error in summary and ensure sheet exists in structured dict
            summary_rows.append([sheet_name, 0, f"ERROR: {e}", 0, 0, 0])
            sheets_structured[sheet_name] = {
                "common_rows": [],
                "different_rows": [],
                "only_in_original_rows": [],
                "only_in_website_rows": [],
                "rows_compared": 0,
                "cells_different": 0,
                "error": str(e)
            }

    # === Move common rows to separate sheets AND rebuild original sheets without common rows ===
    # We'll reconstruct each original sheet from structured lists (diff + only_orig + only_web),
    # and create "<Sheet> Common Rows" for its common rows.
    for sheet_name, sdata in list(sheets_structured.items()):
        # skip if no structured info
        if not isinstance(sdata, dict):
            continue

        # create common sheet if there are actual common data rows (header + >0 rows)
        common_rows = sdata.get("common_rows") or []
        if len(common_rows) > 1:
            ws_common = output_wb.create_sheet(title=f"{sheet_name} Common Rows")
            # write header + data rows
            for r in common_rows:
                ws_common.append(r)

        # Rebuild original sheet to exclude common rows:
        # gather the header (prefer different_rows header, fallback to common/other headers)
        header = None
        for key in ("different_rows", "common_rows", "only_in_original_rows", "only_in_website_rows"):
            lst = sdata.get(key) or []
            if len(lst) > 0:
                header = lst[0]
                break

        # Collect remaining rows (exclude header element)
        remaining_rows = []
        for key in ("different_rows", "only_in_original_rows", "only_in_website_rows"):
            lst = sdata.get(key) or []
            if len(lst) > 1:
                remaining_rows.extend(lst[1:])

        # Replace original sheet with new sheet that contains header + remaining rows
        # Delete the old sheet if present
        if sheet_name in output_wb.sheetnames:
            try:
                del output_wb[sheet_name]
            except Exception:
                pass

        new_ws = output_wb.create_sheet(title=sheet_name)
        if header:
            new_ws.append(header)
        for r in remaining_rows:
            new_ws.append(r)
        # best-effort autosize for the new sheet
        for col_idx in range(1, new_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in new_ws[col_letter]:
                if cell.value is not None:
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        max_len = max(max_len, 10)
            new_ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # overall Summary sheet in the workbook
    ws_summary = output_wb.create_sheet("Summary", index=0)
    ws_summary.append(["Excel Comparison Report"])
    ws_summary.append([f"Generated On:  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    for row in summary_rows:
        ws_summary.append(row)

    out_bytes = io.BytesIO()
    output_wb.save(out_bytes)
    out_bytes.seek(0)

    return {
        "excel_bytes": out_bytes.getvalue(),
        "summary_rows": summary_rows,
        "sheets": sheets_structured
    }

# ---------- End compare function ----------

# Helper that was used previously (kept for compatibility but not used by new flow)
def write_html_report(summary_rows):
    html = """
    <html>
    <head>
        <title>Excel Comparison Report</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 20px; }
            table { border-collapse: collapse; width: 100%; }
            th, td { border: 1px solid #ccc; padding: 6px; text-align: left; }
            th { background: #f2f2f2; }
        </style>
    </head>
    <body>
        <h1>Excel Comparison Report</h1>
        <table>
            <tr>
                <th>Sheet</th>
                <th>Rows Compared</th>
                <th>Cells Different</th>
                <th>Common Rows</th>
                <th>Only in Original</th>
                <th>Only in Website</th>
            </tr>
    """
    for row in summary_rows[1:]:  # skip header row
        html += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"

    html += """
        </table>
    </body>
    </html>
    """
    return html

# ---------- API endpoints (updated to use the structured sheets) ----------

@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.utcnow().isoformat() + "Z"}

# Download full workbook
@app.post("/compare")
async def compare_endpoint(
    original_file: UploadFile = File(...),
    website_file: UploadFile = File(...),
):
    original_bytes = await original_file.read()
    website_bytes  = await website_file.read()

    result = compare_excel_with_gain_summary_inline(original_bytes, website_bytes)

    filename = f"comparison_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        io.BytesIO(result["excel_bytes"]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# JSON summary for the summary table (keeps old shape for backward compatibility)
@app.post("/compare/json")
async def compare_json(
    original_file: UploadFile = File(...),
    website_file: UploadFile = File(...),
    sheets_config: Optional[str] = Form(default=None),
):
    try:
        parsed_config = None
        if sheets_config and sheets_config.strip():
            try:
                parsed_config = json.loads(sheets_config.strip())
            except Exception:
                parsed_config = None

        original_bytes = await original_file.read()
        website_bytes  = await website_file.read()

        result = compare_excel_with_gain_summary_inline(
            original_bytes, website_bytes, sheets_config=parsed_config
        )

        # Return the same "summary_rows" array (for your index.html summary rendering)
        return JSONResponse(content={"results": result["summary_rows"]})
    except Exception as e:
        return {"error": str(e)}

# HTML report (optional)
@app.post("/compare/html")
async def compare_html(
    original_file: UploadFile = File(...),
    website_file: UploadFile = File(...),
    sheets_config: Optional[str] = Form(default=None),
):
    try:
        parsed_config = None
        if sheets_config and sheets_config.strip():
            try:
                parsed_config = json.loads(sheets_config.strip())
            except Exception:
                parsed_config = None

        original_bytes = await original_file.read()
        website_bytes  = await website_file.read()

        result = compare_excel_with_gain_summary_inline(
            original_bytes, website_bytes, sheets_config=parsed_config
        )

        html = write_html_report(result["summary_rows"])
        return HTMLResponse(content=html, media_type="text/html")
    except Exception as e:
        return {"error": str(e)}

# Return list of sheets that have any differences (uses structured sheets, not color scanning)
@app.post("/compare/sheets")
async def compare_sheets(
    original_file: UploadFile = File(...),
    website_file: UploadFile = File(...),
    sheets_config: Optional[str] = Form(default=None)
):
    parsed_config = None
    if sheets_config:
        try:
            parsed_config = json.loads(sheets_config)
        except:
            parsed_config = None

    original_bytes = await original_file.read()
    website_bytes  = await website_file.read()

    result = compare_excel_with_gain_summary_inline(original_bytes, website_bytes, sheets_config=parsed_config)

    sheets_with_diff = []
    for sname, sdata in result["sheets"].items():
        if (len(sdata.get("different_rows", [])) > 1 or
            len(sdata.get("only_in_original_rows", [])) > 1 or
            len(sdata.get("only_in_website_rows", [])) > 1):
            # lists include a header row, so >1 indicates actual data rows
            sheets_with_diff.append(sname)

    return {"sheets": sheets_with_diff}

# Return categorized rows for a specific sheet (structured)
@app.post("/compare/sheet/diff")
async def compare_sheet_diff(
    original_file: UploadFile = File(...),
    website_file: UploadFile = File(...),
    sheet_name: str = Form(...),
    sheets_config: Optional[str] = Form(default=None)
):
    parsed_config = None
    if sheets_config:
        try:
            parsed_config = json.loads(sheets_config)
        except:
            parsed_config = None

    original_bytes = await original_file.read()
    website_bytes  = await website_file.read()

    result = compare_excel_with_gain_summary_inline(
        original_bytes,
        website_bytes,
        sheets_config=parsed_config
    )

    sheets = result.get("sheets", {})
    target = sheets.get(sheet_name)
    if not target:
        # if missing, return empty sets (keeps UI stable)
        return {
            "sheet_name": sheet_name,
            "common_rows": [],
            "different_rows": [],
            "only_in_original_rows": [],
            "only_in_website_rows": []
        }

    return {
        "sheet_name": sheet_name,
        "common_rows": target.get("common_rows", []),
        "different_rows": target.get("different_rows", []),
        "only_in_original_rows": target.get("only_in_original_rows", []),
        "only_in_website_rows": target.get("only_in_website_rows", [])
    }

from fastapi.responses import FileResponse
import tempfile
import os

@app.post("/compare/excel")
async def compare_excel(original_file: UploadFile = File(...), website_file: UploadFile = File(...)):
    # Save uploaded files temporarily
    temp_dir = tempfile.mkdtemp()
    orig_path = os.path.join(temp_dir, original_file.filename)
    web_path  = os.path.join(temp_dir, website_file.filename)

    with open(orig_path, "wb") as f:
        f.write(await original_file.read())
    with open(web_path, "wb") as f:
        f.write(await website_file.read())

    # Call the comparison function and write the result to the output file
    output_path = os.path.join(temp_dir, "Comparison_Result.xlsx")
    with open(orig_path, "rb") as f1, open(web_path, "rb") as f2:
        original_bytes = f1.read()
        website_bytes = f2.read()
    result = compare_excel_with_gain_summary_inline(original_bytes, website_bytes)
    with open(output_path, "wb") as out_f:
        out_f.write(result["excel_bytes"])

    return FileResponse(output_path, filename="Comparison_Result.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
